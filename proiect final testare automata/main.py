import openpyxl
from openpyxl.chart import(PieChart, Reference)
from tkinter import *
import win32com.client as win32

root = Tk()
root.geometry("400x400")
root.resizable(0,0)
root.title("Test Cases parser")
Label(root, text = "Test Cases Parser", font ="arial 15 bold").pack()

values=[0,0]
passString = "Pass"
failString = "Fail"
path = r"C:\Users\gheor\OneDrive\Desktop\Proiect final Ioana\Test_Case_Format_(Ioana Stoinea).xlsx"
pathPDF = r"C:\Users\gheor\OneDrive\Desktop\Proiect final Ioana\Test_Case_Format_(Ioana Stoinea).pdf"

tester_label = Label(root, text = "Tester Name", font ="arial 15 bold").pack()
tester_str = StringVar
Entry(root, textvariable=tester_str).pack()

def generateReport():
    wb = openpyxl.load_workbook(path, read_only=False)
    testCasesSheet = wb.get_sheet_by_name('TestCase')
    global tester
    tester = testCasesSheet['E1'].value
    try:
        reportSheet = wb.get_sheet_by_name('Report')

    except:
        wb.create_sheet("Report")
        reportSheet = wb.get_sheet_by_name('Report')
    reportSheet['A1'] = "TesterID: "
    reportSheet['B1'] = tester
    reportSheet['A2'] = "Failed test cases "
    reportSheet["B2"] = values[1]
    reportSheet["A3"] = "Passed test cases "
    reportSheet["B3"] = values[0]
    reportSheet["A4"] = "Total number of test cases "
    reportSheet["B4"] = values[0] + values[1]
    wb.save(path)
    createChart()

    # Open Excel File
    excel = win32.Dispatch("Excel.Application")
    #Read excel file
    sheets = excel.Workbooks.Open(path)
    work_sheets = sheets.Worksheets[2]
    # convert into PDF file
    work_sheets.ExportAsFixedFormat(0, pathPDF)

def compareValues():
    wb = openpyxl.load_workbook(path, read_only=False)
    testCasesSheet = wb.get_sheet_by_name('TestCase')

    for row in range(1, int(testCasesSheet.max_row +1 )):
        if testCasesSheet.cell(row=row, column=7).value == "Pass":
           values[0]=values[0]+1
        elif testCasesSheet.cell(row=row, column=7).value == "Fail":
            values[1]=values[1]+1

def createChart():
    wb = openpyxl.load_workbook(path, read_only=False)
    sheet = wb["Report"]
    pie = PieChart()

    labels = Reference(sheet, min_col=1, min_row=2, max_row=3)
    data = Reference(sheet, min_col=2, min_row=2, max_row=3)
    pie.add_data(data, titles_from_data = False)
    pie.set_categories(labels)
    pie.title = "Test Cases"

    pie.width = 14
    pie.height = 7

    s=pie.series[0]

    s.graphicalProperties.line.solodFill = "00000"

    sheet.add_chart(pie, "A6")

    wb.save(path)


def buttonPressed():
    compareValues()
    generateReport()

Button(root, text="Generate Report", command = buttonPressed).pack(pady=5)

root.mainloop()





